#!/usr/bin/env python3

import argparse
import csv
import re
import sys
import zipfile
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

W_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
A_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
R_NS = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
SS_NS = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

TEXT_SUFFIXES = {".txt", ".md"}
DELIMITED_SUFFIXES = {".csv": ",", ".tsv": "\t"}
SUPPORTED_SUFFIXES = TEXT_SUFFIXES | set(DELIMITED_SUFFIXES) | {".docx", ".pptx", ".xlsx"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract text locally from common document formats.")
    parser.add_argument("file_path", help="Path to a local document")
    parser.add_argument("--max-chars", type=int, default=120000, help="Maximum number of output characters")
    parser.add_argument("--max-rows-per-sheet", type=int, default=200, help="Maximum number of rows per sheet")
    return parser.parse_args()


def resolve_input(path_arg: str) -> Path:
    path = Path(path_arg).expanduser().resolve()
    if not path.exists():
        raise RuntimeError(f"document not found: {path}")
    if not path.is_file():
        raise RuntimeError(f"path is not a file: {path}")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        supported = ", ".join(sorted(SUPPORTED_SUFFIXES))
        raise RuntimeError(f"unsupported file type {suffix or '<none>'}; supported: {supported}")
    return path


def decode_text_bytes(data: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "utf-16", "utf-16-le", "utf-16-be"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1")


def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def truncate_output(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    clipped = text[:max_chars].rstrip()
    return clipped + "\n\n[TRUNCATED: output limit reached]"


def parse_text_file(path: Path) -> str:
    return normalize_text(decode_text_bytes(path.read_bytes()))


def parse_delimited_file(path: Path, delimiter: str) -> str:
    content = decode_text_bytes(path.read_bytes())
    rows = []
    reader = csv.reader(content.splitlines(), delimiter=delimiter)
    for index, row in enumerate(reader, start=1):
        cells = [cell.strip() for cell in row]
        while cells and cells[-1] == "":
            cells.pop()
        rows.append(f"{index}\t" + "\t".join(cells))
    return normalize_text("\n".join(rows))


def parse_docx(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        try:
            document_xml = archive.read("word/document.xml")
        except KeyError as exc:
            raise RuntimeError("invalid docx: missing word/document.xml") from exc

    root = ET.fromstring(document_xml)
    body = root.find("w:body", W_NS)
    if body is None:
        raise RuntimeError("invalid docx: missing document body")

    sections: list[str] = []
    for child in body:
        local_name = child.tag.rsplit("}", 1)[-1]
        if local_name == "p":
            text = "".join(node.text or "" for node in child.findall(".//w:t", W_NS)).strip()
            if text:
                sections.append(text)
        elif local_name == "tbl":
            rows = []
            for row in child.findall("w:tr", W_NS):
                cells = []
                for cell in row.findall("w:tc", W_NS):
                    cell_text = " ".join(
                        filter(
                            None,
                            ("".join(node.text or "" for node in para.findall(".//w:t", W_NS)).strip()
                             for para in cell.findall(".//w:p", W_NS)),
                        )
                    ).strip()
                    cells.append(cell_text)
                while cells and cells[-1] == "":
                    cells.pop()
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sections.append("[Table]\n" + "\n".join(rows))

    text = normalize_text("\n\n".join(sections))
    if not text:
        raise RuntimeError("no extractable text found in docx")
    return text


def slide_sort_key(name: str) -> int:
    match = re.search(r"slide(\d+)\.xml$", name)
    return int(match.group(1)) if match else 0


def parse_pptx(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        slide_names = sorted(
            (name for name in archive.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)),
            key=slide_sort_key,
        )
        if not slide_names:
            raise RuntimeError("invalid pptx: no slide XML found")

        sections = []
        for index, slide_name in enumerate(slide_names, start=1):
            root = ET.fromstring(archive.read(slide_name))
            paragraphs = []
            for paragraph in root.findall(".//a:p", A_NS):
                text = "".join(node.text or "" for node in paragraph.findall(".//a:t", A_NS)).strip()
                if text:
                    paragraphs.append(text)
            slide_text = "\n".join(paragraphs).strip() or "[No extractable text on this slide]"
            sections.append(f"[Slide {index}]\n{slide_text}")

    text = normalize_text("\n\n".join(sections))
    if "[No extractable text on this slide]" in text and len(sections) == 1 and text.strip() == "[Slide 1]\n[No extractable text on this slide]":
        raise RuntimeError("no extractable text found in pptx")
    return text


def stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def parse_xlsx(path: Path, max_rows_per_sheet: int) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("missing dependency: install openpyxl to parse xlsx files locally") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sections = []
        for sheet in workbook.worksheets:
            rows = []
            emitted_rows = 0
            truncated = False
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [stringify_cell(value) for value in row]
                while cells and cells[-1] == "":
                    cells.pop()
                if not any(cells):
                    continue
                rows.append(f"{row_index}\t" + "\t".join(cells))
                emitted_rows += 1
                if emitted_rows >= max_rows_per_sheet:
                    truncated = True
                    break

            body = "\n".join(rows).strip() or "[No non-empty rows]"
            if truncated:
                body += "\n[TRUNCATED: sheet row limit reached]"
            sections.append(f"[Sheet: {sheet.title}]\n{body}")
    finally:
        workbook.close()

    text = normalize_text("\n\n".join(sections))
    if not text:
        raise RuntimeError("no extractable content found in xlsx")
    return text


def extract_document(path: Path, max_rows_per_sheet: int) -> str:
    suffix = path.suffix.lower()
    if suffix in TEXT_SUFFIXES:
        text = parse_text_file(path)
    elif suffix in DELIMITED_SUFFIXES:
        text = parse_delimited_file(path, DELIMITED_SUFFIXES[suffix])
    elif suffix == ".docx":
        text = parse_docx(path)
    elif suffix == ".pptx":
        text = parse_pptx(path)
    elif suffix == ".xlsx":
        text = parse_xlsx(path, max_rows_per_sheet=max_rows_per_sheet)
    else:
        raise RuntimeError(f"unsupported file type {suffix or '<none>'}")

    if not text:
        raise RuntimeError(f"no readable content found in {path.name}")
    return text


def main() -> int:
    args = parse_args()

    try:
        document_path = resolve_input(args.file_path)
        text = extract_document(document_path, max_rows_per_sheet=max(args.max_rows_per_sheet, 1))
        text = truncate_output(text, max(args.max_chars, 1))
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1

    print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
